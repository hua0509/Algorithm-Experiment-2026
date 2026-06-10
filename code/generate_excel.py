# -*- coding: utf-8 -*-
"""
生成Excel数据文件 - 0-1背包1000个物品的完整数据
"""
import random
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE_DIR = r'E:\算法设计与分析实验'
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

def generate_excel():
    wb = Workbook()
    
    # ============ Sheet 1: 1000个物品数据 ============
    ws1 = wb.active
    ws1.title = '1000个物品数据(C=10000)'
    
    # 表头样式
    header_font = Font(name='宋体', size=11, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ['物品编号', '物品重量', '物品价值']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # 生成数据
    random.seed(42 + 1000)  # seed = 1042
    data_font = Font(name='Times New Roman', size=10)
    
    for i in range(1000):
        w = random.randint(1, 100)
        v = round(random.uniform(100, 1000), 2)
        
        ws1.cell(row=i+2, column=1, value=i+1).font = data_font
        ws1.cell(row=i+2, column=2, value=w).font = data_font
        ws1.cell(row=i+2, column=3, value=v).font = data_font
        
        for col in [1, 2, 3]:
            ws1.cell(row=i+2, column=col).border = thin_border
            ws1.cell(row=i+2, column=col).alignment = Alignment(horizontal='center')
    
    # 调整列宽
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15
    
    # ============ Sheet 2: 排序算法数据 ============
    ws2 = wb.create_sheet('排序算法比较次数')
    
    sort_headers = ['输入规模n', '冒泡排序比较次数', '合并排序比较次数', '快速排序比较次数']
    for col, h in enumerate(sort_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    sort_data = [
        [10, 45, 23, 29],
        [100, 4697, 545, 614],
        [1000, 499329, 8705, 11972],
        [2000, 1997824, 19436, 26157],
        [5000, 12492247, 55230, 78890],
        [10000, 49994847, 120485, 154959],
        [100000, 4999950000, 1536229, 2070153],
    ]
    
    for i, row in enumerate(sort_data):
        for j, val in enumerate(row):
            cell = ws2.cell(row=i+2, column=j+1, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 22
    
    # ============ Sheet 3: 0-1背包DP+贪心执行时间 ============
    ws3 = wb.create_sheet('背包问题执行时间')
    
    time_headers = ['物品数量n', 'C=10000 DP(ms)', 'C=10000 贪心(ms)', 
                    'C=100000 DP(ms)', 'C=100000 贪心(ms)',
                    'C=1000000 贪心(ms)']
    for col, h in enumerate(time_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    import json
    with open(r'C:\Users\41529\WorkBuddy\2026-06-10-16-03-30\experiment_data\knapsack_results.json', 'r', encoding='utf-8') as f:
        kr = json.load(f)
    
    n_list = [1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 
              10000, 20000, 40000, 80000, 160000, 320000]
    
    for i, n in enumerate(n_list):
        row_data = [n]
        for cap_str in ['10000', '100000', '1000000']:
            r = kr[cap_str].get(str(n), {})
            dp_t = r.get('dp', {}).get('execution_time_ms', '-')
            gd_t = r.get('greedy', {}).get('execution_time_ms', '-')
            if isinstance(dp_t, (int, float)):
                row_data.append(round(dp_t, 2))
            else:
                row_data.append('-')
            row_data.append(round(gd_t, 2))
        
        # Remove duplicate greedy for last capacity
        row_data_final = row_data[:6]  # n, dp10k, greedy10k, dp100k, greedy100k, greedy1M
        
        for j, val in enumerate([n,
                                 row_data[1], row_data[2],
                                 row_data[3], row_data[4],
                                 row_data[5]]):
            cell = ws3.cell(row=i+2, column=j+1, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col_letter].width = 18
    
    # 保存
    excel_path = os.path.join(DATA_DIR, '20241120062-刘兴华-数据.xlsx')
    wb.save(excel_path)
    print(f'Excel saved to: {excel_path}')
    return excel_path

if __name__ == '__main__':
    generate_excel()
